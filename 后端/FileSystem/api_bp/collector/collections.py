import os
import shutil
import time
import zipfile
from datetime import datetime

from openpyxl import Workbook
from flask import Blueprint, jsonify, request, send_file, make_response
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from configs import PROJECT_URL, COLLECTIONS_FOLDER_URL
from data_base import Class, ClassStudentMap, User, Collection, CollectionItems, CollectionRecord, remind_email
from utils.JWT import auth, get_id, get_user_roles, get_token_from_request
from utils.my_email import zip_email

collections_bp=Blueprint('collections',__name__)

def create_excel(path,collection_id):
    collating_record = CollectionRecord.collating_record(collection_id)
    collection = Collection.get_collection(collection_id)
    items_info = CollectionItems.get_collection_items_info(collection_id)
    table_name=f"{collection.class_name}  {collection.collection_name}  收集情况表"

    wb = Workbook()
    ws = wb.active
    #标题
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1, value=table_name)
    title.font = Font(name=u'宋体', size=16, bold=True, color='205DCD')
    title.alignment = Alignment(horizontal='center')
    title.border = Border(bottom=Side(border_style='thin'))
    title.fill = PatternFill(fgColor='FFFFAB', fill_type='solid')
    #字段(第二行)
    columns = ['学号', '姓名', '提交数量']
    columns.extend(items_info)
    ws.append(columns)

    for record in collating_record:
        row = list(record[:3])
        little_record = [int(lr) for lr in record[-1].split(',')]  # 有交的项的index列表
        for index in range(collection.collection_items_amount):
            row.append('Y' if index in little_record else 'N')
        ws.append(row)
    ws.auto_filter.ref = f"A2:{chr(ord('A') + len(columns) - 1)}2"
    for row in ws[f"C3:{chr(ord('A') + len(columns) - 1)}{ws.max_column}"]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    # 调整列宽
    ws.column_dimensions['A'].width = 11
    wb.save(f"{path}{table_name}.xlsx")
    return f"{path}{table_name}.xlsx"
def zip_collection(collection_id):
    # https://blog.csdn.net/VinWqx/article/details/108842701   单个目录压缩优化2
    collection = Collection.get_collection(collection_id)
    path = COLLECTIONS_FOLDER_URL + collection.collection_id + '/'
    zip_target_path = path
    zip_result_path = COLLECTIONS_FOLDER_URL + 'packages/' + collection.collection_id + '.zip'

    if os.path.exists(zip_target_path):
        create_excel(path, collection.collection_id)
        zip = zipfile.ZipFile(zip_result_path, 'w', zipfile.ZIP_DEFLATED)
        for dir_path, dir_names, file_names in os.walk(zip_target_path):
            fpath = dir_path.replace(zip_target_path, '')
            for filename in file_names:
                zip.write(os.path.join(dir_path, filename), os.path.join(fpath, filename))
        zip.close()
    return {'file_path': zip_result_path, 'file_name': collection.class_name+'  '+collection.collection_name,'class_name':collection.class_name,'collection_name':collection.collection_name}


@collections_bp.route('/new_collection', methods=['POST'])
@auth.login_required(role='collector')
def new_collection():
    form = request.form
    print(form)
    form['collection_items']=eval(form['collection_items'].replace('true','True'))
    form['collection_items_amount']=len(form['collection_items'])
    form['collection_start_time'] = datetime.fromtimestamp(form['collection_start_time']/1000)
    form['collection_end_time'] = datetime.fromtimestamp(form['collection_end_time']/1000)
    collection_id = Collection.new_collection(form)
    path = COLLECTIONS_FOLDER_URL + collection_id + '/'
    if not os.path.exists(path):  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)
    for i in range(len(form['collection_items'])):
        collection_item=form['collection_items'][i]
        item={}
        item['collection_id'] = collection_id
        item['index'] = i
        item['info'] = collection_item['info']
        item['type'] = collection_item['type']
        # item['necessary'] = collection_item['necessary']
        CollectionItems.new_collection_item(item)
        collection_path=path+str(item['index'])+'_'+item['info']+ '/'
        if not os.path.exists(collection_path):  # 判断是否存在文件夹如果不存在则创建为文件夹
            os.makedirs(collection_path)
    return '',200

@collections_bp.route('/delete_collection', methods=['POST'])
@auth.login_required(role='collector')
def delete_collection():
    Collection.delete_collection(request.form['collection_id'])
    shutil.rmtree(COLLECTIONS_FOLDER_URL + request.form['collection_id'] + '/')
    return '', 200

@collections_bp.route('/update_collection', methods=['POST'])
@auth.login_required(role='collector')
def update_collection():
    form = request.form
    if form['field']=='collection_start_time' or form['field']=='collection_end_time':
        form['data'] = datetime.fromtimestamp(form['data']/1000)
    Collection.update(form['collection_id'],form['field'],form['data'])
    return '', 200


@collections_bp.route('/email_collection_zip', methods=['POST'])
@auth.login_required(role='collector')
def email_collection_zip():
    zip_file = zip_collection(request.form['collection_id'])
    zip_email(User.get_user(get_id()).email,zip_file)
    return '',200
@collections_bp.route('/download_collection_zip', methods=['POST'])
@auth.login_required(role='collector')
def download_collection_zip():
    # https://blog.csdn.net/VinWqx/article/details/108842701   单个目录压缩优化2
    collection = Collection.get_collection(request.form['collection_id'])
    zip_file = zip_collection(collection.collection_id)
    response = make_response(send_file(zip_file['file_path'], as_attachment=True))
    response.headers["Content-Disposition"] = f"attachment; filename={zip_file['file_name'].encode().decode('latin-1')}"
    response.direct_passthrough = False
    print(response.data.hex()[:32])
    return response

@collections_bp.route('/send_remind_email', methods=['POST'])
@auth.login_required(role='collector')
def send_remind_email():
    remind_email(Collection.get_collection(request.form['collection_id']))
    return '', 200
